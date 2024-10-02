import os
from openpyxl import load_workbook, Workbook
import logging

logging.basicConfig(filename='zalohy.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

class ZalohyManager:
    def __init__(self):
        self.excel_cesta = "Hodiny_Cap.xlsx"
        self.ZALOHY_SHEET_NAME = 'Zálohy'
        self.EMPLOYEE_START_ROW = 9

    def nacti_nebo_vytvor_excel(self):
        try:
            if os.path.exists(self.excel_cesta):
                workbook = load_workbook(self.excel_cesta)
                logging.info(f"Načten existující Excel soubor: {self.excel_cesta}")
            else:
                workbook = Workbook()
                workbook.save(self.excel_cesta)
                logging.info(f"Vytvořen nový Excel soubor: {self.excel_cesta}")
            
            if self.ZALOHY_SHEET_NAME not in workbook.sheetnames:
                workbook.create_sheet(self.ZALOHY_SHEET_NAME)
                logging.info(f"Vytvořen nový list '{self.ZALOHY_SHEET_NAME}'")
            
            return workbook
        except Exception as e:
            logging.error(f"Chyba při načítání nebo vytváření Excel souboru: {e}")
            raise

    def get_employee_row(self, employee_name):
        workbook = self.nacti_nebo_vytvor_excel()
        sheet = workbook[self.ZALOHY_SHEET_NAME]
        for row in range(self.EMPLOYEE_START_ROW, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == employee_name:
                return row
        return None

    def add_employee_advance(self, employee_name, amount, currency):
        workbook = self.nacti_nebo_vytvor_excel()
        sheet = workbook[self.ZALOHY_SHEET_NAME]
        row = self.get_employee_row(employee_name)
        
        if row is None:
            row = self.get_next_empty_row(sheet)
            sheet.cell(row=row, column=1, value=employee_name)
        
        if currency == '€':
            current_value = sheet.cell(row=row, column=2).value or 0
            sheet.cell(row=row, column=2, value=current_value + amount)
        else:
            current_value = sheet.cell(row=row, column=3).value or 0
            sheet.cell(row=row, column=3, value=current_value + amount)
        
        workbook.save(self.excel_cesta)
        logging.info(f"Záloha pro {employee_name} aktualizována: {amount} {currency}")
        
    def update_employee_advance(self, row, amount, currency):
        """Updates the advance amount for an existing employee."""
        workbook = self.nacti_nebo_vytvor_excel()
        sheet = workbook[self.ZALOHY_SHEET_NAME]
        
        if currency == '€':
            current_value = sheet[f'B{row}'].value
            if current_value is None:
                sheet[f'B{row}'] = amount
            else:
                sheet[f'B{row}'] = current_value + amount
        else:
            current_value = sheet[f'C{row}'].value
            if current_value is None:
                sheet[f'C{row}'] = amount
            else:
                sheet[f'C{row}'] = current_value + amount
        
        workbook.save(self.excel_cesta)

    def get_next_empty_row(self, sheet):
        for row in range(self.EMPLOYEE_START_ROW, sheet.max_row + 2):
            if sheet.cell(row=row, column=1).value is None:
                return row
        return sheet.max_row + 1

    def get_employee_advances(self, employee_name):
        workbook = self.nacti_nebo_vytvor_excel()
        sheet = workbook[self.ZALOHY_SHEET_NAME]
        row = self.get_employee_row(employee_name)
        if row is None:
            return None
        return {
            'EUR': sheet.cell(row=row, column=2).value or 0,
            'CZK': sheet.cell(row=row, column=3).value or 0
        }

if __name__ == "__main__":
    # Test code
    manager = ZalohyManager()
    manager.add_or_update_employee_advance("Jan Novák", 100, '€')
    manager.add_or_update_employee_advance("Jan Novák", 2000, 'Kč')
    print(manager.get_employee_advances("Jan Novák"))
