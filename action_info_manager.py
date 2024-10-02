import openpyxl
from datetime import datetime
import logging

logging.basicConfig(filename='action_info.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

class ActionInfoManager:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def write_action_info(self, action_name, start_date, end_date):
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            
            if 'Zálohy' not in workbook.sheetnames:
                workbook.create_sheet('Zálohy')
            
            sheet = workbook['Zálohy']
            
            sheet['B4'] = f"NÁZEV PROJEKTU: ({action_name})"
            
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            date_range = f"{start_date.strftime('%d.%m.%y')} - {end_date.strftime('%d.%m.%y')}"
            sheet['B3'] = date_range
            
            workbook.save(self.excel_path)
            logging.info(f"Informace o akci '{action_name}' byly úspěšně zapsány do Excel souboru.")
        except Exception as e:
            logging.error(f"Chyba při zápisu informací o akci: {str(e)}")
            raise

    def clear_action_info(self):
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            
            if 'Zálohy' in workbook.sheetnames:
                sheet = workbook['Zálohy']
                sheet['B3'] = ""
                sheet['B4'] = ""
                
                workbook.save(self.excel_path)
                logging.info("Informace o akci byly úspěšně vymazány z Excel souboru.")
        except Exception as e:
            logging.error(f"Chyba při mazání informací o akci: {str(e)}")
            raise